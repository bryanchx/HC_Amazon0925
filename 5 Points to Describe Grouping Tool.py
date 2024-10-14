import pandas as pd
import random
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import openpyxl
from openpyxl import load_workbook

def split_into_groups(data, target_length=465):
    groups = []
    current_group = []
    current_length = 0

    random.shuffle(data)  # 随机打乱数据

    for item in data:
        item_length = len(item)
        # 如果添加当前项后超出目标长度，则创建新组
        if current_length + item_length > target_length:
            if current_group:  # 确保当前组非空
                groups.append(current_group)
            current_group = [item]
            current_length = item_length
        else:
            current_group.append(item)
            current_length += item_length

    # 添加最后一组
    if current_group:
        groups.append(current_group)

    # 检查最后一组的长度，如果超出目标长度则拆分
    final_groups = []
    for group in groups:
        group_string = ' '.join(group)
        if len(group_string) > target_length:
            split_groups = []
            temp_group = []
            temp_length = 0
            for item in group:
                item_length = len(item)
                if temp_length + item_length > target_length:
                    split_groups.append(temp_group)
                    temp_group = [item]
                    temp_length = item_length
                else:
                    temp_group.append(item)
                    temp_length += item_length
            if temp_group:  # 添加剩余的部分
                split_groups.append(temp_group)
            final_groups.extend(split_groups)
        else:
            final_groups.append(group)

    return final_groups

def main():
    Tk().withdraw()
    file_path = askopenfilename(title="选择Excel文件", filetypes=[("Excel files", "*.xlsx;*.xls")])

    if not file_path:
        print("未选择文件，程序结束。")
        return

    df = pd.read_excel(file_path)
    data_column = df.iloc[:, 0].dropna().tolist()

    try:
        num_runs = int(input("请输入要执行的次数: "))
    except ValueError:
        print("输入无效，请输入一个整数。")
        return

    wb = load_workbook(file_path)

    if 'Grouped Results' in wb.sheetnames:
        sheet = wb['Grouped Results']
    else:
        sheet = wb.create_sheet(title='Grouped Results')

    next_row = sheet.max_row + 1

    for run in range(num_runs):
        random_data = data_column[:]  # 创建数据的副本以便每次随机选择
        random.shuffle(random_data)  # 随机打乱数据

        grouped_data = split_into_groups(random_data)
        for i, group in enumerate(grouped_data):
            group_string = ' '.join(group)  # 将组内元素连接成字符串
            sheet.cell(row=next_row, column=i + 1, value=group_string)

        next_row += 1  # 更新下一行的位置

    wb.save(file_path)
    print(f'分组结果已成功写入到 "{file_path}" 的 Sheet "Grouped Results" 中.')

if __name__ == "__main__":
    main()
