import pandas as pd
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from datetime import datetime

# 我有一个需求：
# 1、手动选择excel，并且列出sheet让我选择；
# 2、读取sheet每列数据，把里面的所有单词首字母大写，介词除外；
# 3、新的单词组存储在新的sheet，用时间戳排序。

# 创建隐藏的Tkinter窗口
Tk().withdraw()

# 手动选择Excel文件
file_path = askopenfilename(
    title="选择Excel文件",
    filetypes=[("Excel files", "*.xlsx;*.xls")]
)

if not file_path:
    print("未选择文件，程序结束。")
    exit()

# 加载Excel工作簿
wb = load_workbook(file_path)

# 获取所有Sheet名称
sheet_names = wb.sheetnames

# 打印所有Sheet名称，供用户选择
print("可用的Sheet:")
for i in range(len(sheet_names)):
    name = sheet_names[i]
    print(f"{i + 1}: {name}")

# 用户选择Sheet
try:
    sheet_index = int(input("请输入要修改的Sheet编号: ")) - 1
    if sheet_index < 0 or sheet_index >= len(sheet_names):
        raise ValueError("选择的编号无效。")
except ValueError as e:
    print(f"输入无效: {e}")
    exit()

# 读取选定的Sheet
selected_sheet_name = sheet_names[sheet_index]
df = pd.read_excel(file_path, sheet_name=selected_sheet_name)

# 定义函数用于将单词首字母大写
def capitalize_title(title):
    prepositions = {"and", "or", "in", "on", "at", "for", "with", "but", "by", "the", "a", "an", "to", "of"}
    words = title.split()
    capitalized_words = []

    for word in words:
        if word.lower() in prepositions:
            capitalized_words.append(word.lower())
        else:
            capitalized_words.append(word.capitalize())

    return ' '.join(capitalized_words)

# 创建一个新的DataFrame以存储修改后的内容
new_df = pd.DataFrame()

# 处理每一列
for col in df.columns:
    new_column_data = df[col].astype(str).apply(capitalize_title)
    new_df[col] = new_column_data

# 生成时间戳
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
new_sheet_name = f'Modified_{selected_sheet_name}_{timestamp}'

# 打开现有的Excel工作簿并保存新Sheet
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
    new_df.to_excel(writer, sheet_name=new_sheet_name, index=False)

print(f"新数据已成功存储至新的Sheet: '{new_sheet_name}'。")
