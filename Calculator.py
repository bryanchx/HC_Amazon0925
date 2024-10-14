import pandas as pd
from openpyxl import load_workbook
from tkinter import Tk, Button, Label, filedialog, StringVar, Listbox, END, Frame
from datetime import datetime

class ExcelProcessor:
    def __init__(self, master):
        self.master = master
        master.title("Excel Processor")

        self.label = Label(master, text="选择Excel文件:")
        self.label.pack()

        self.select_button = Button(master, text="选择文件", command=self.select_file)
        self.select_button.pack()

        self.sheet_label = Label(master, text="可用的Sheet:")
        self.sheet_label.pack()

        self.sheet_listbox = Listbox(master)
        self.sheet_listbox.pack()

        self.capitalize_button = Button(master, text="处理首字母大写", command=self.capitalize_titles)
        self.capitalize_button.pack()

        self.group_button = Button(master, text="分组处理", command=self.group_data)
        self.group_button.pack()

        self.message = StringVar()
        self.message_label = Label(master, textvariable=self.message)
        self.message_label.pack()

        self.file_path = None
        self.sheet_names = []

    def select_file(self):
        self.file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if self.file_path:
            self.load_sheets()

    def load_sheets(self):
        workbook = load_workbook(self.file_path)
        self.sheet_names = workbook.sheetnames
        self.sheet_listbox.delete(0, END)  # 清空现有内容
        for sheet in self.sheet_names:
            self.sheet_listbox.insert(END, sheet)

    def capitalize_title(self, title):
        prepositions = {"and", "or", "in", "on", "at", "for", "with", "but", "by", "the", "a", "an", "to", "of"}
        words = title.split()
        capitalized_words = []
        for word in words:
            if word.lower() in prepositions:
                capitalized_words.append(word.lower())
            else:
                capitalized_words.append(word.capitalize())
        return ' '.join(capitalized_words)

    def capitalize_titles(self):
        selected_index = self.sheet_listbox.curselection()
        if not selected_index:
            self.message.set("请先选择一个Sheet！")
            return

        selected_sheet_name = self.sheet_names[selected_index[0]]
        df = pd.read_excel(self.file_path, sheet_name=selected_sheet_name)

        new_df = pd.DataFrame()
        for col in df.columns:
            new_column_data = df[col].astype(str).apply(self.capitalize_title)
            new_df[col] = new_column_data

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_sheet_name = f'Modified_{selected_sheet_name}_{timestamp}'

        # 限制Sheet名称长度
        if len(new_sheet_name) > 31:
            new_sheet_name = new_sheet_name[:31]

        with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a') as writer:
            new_df.to_excel(writer, sheet_name=new_sheet_name, index=False)

        self.message.set(f"新数据已成功存储至新的Sheet: '{new_sheet_name}'。")

    def group_data(self):
        # 这里可以实现分组处理的逻辑
        # 添加具体的分组代码
        self.message.set("分组处理功能尚未实现。")

if __name__ == "__main__":
    root = Tk()
    excel_processor = ExcelProcessor(root)
    root.mainloop()
