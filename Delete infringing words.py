import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# 创建一个简单的 GUI 让用户选择 Excel 文件
def select_file():
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    file_path = filedialog.askopenfilename(
        title="选择 Excel 文件",
        filetypes=[("Excel files", "*.xlsx;*.xls")]
    )
    return file_path


# 列出工作表及其列
def list_sheets_and_columns(excel_file):
    sheets = pd.ExcelFile(excel_file).sheet_names
    sheet_columns = {}

    for sheet in sheets:
        df = pd.read_excel(excel_file, sheet_name=sheet)
        sheet_columns[sheet] = df.columns.tolist()

    return sheet_columns


# 打印工作表和列的选项
def print_sheets_and_columns(sheets_and_columns):
    print("有数据的工作表及列:")
    sheet_options = list(sheets_and_columns.keys())

    for idx, sheet in enumerate(sheet_options):
        print(f"{idx + 1}. 工作表: {sheet}")
        for col_idx, column in enumerate(sheets_and_columns[sheet]):
            print(f"   列 {col_idx + 1}: {column}")

    return sheet_options


# 主函数
def main():
    # 1. 选择 Excel 文件
    excel_file = select_file()
    if not excel_file:
        print("没有选择文件，程序结束。")
        return

    # 2. 列出所有工作表及其列
    sheets_and_columns = list_sheets_and_columns(excel_file)
    sheet_options = print_sheets_and_columns(sheets_and_columns)

    # 让用户选择工作表
    sheet_index = int(input("请输入要处理的工作表编号: ")) - 1
    if sheet_index < 0 or sheet_index >= len(sheet_options):
        print("所选择的工作表编号无效。")
        return

    selected_sheet = sheet_options[sheet_index]

    # 让用户选择关键词列和目标列
    column_names = sheets_and_columns[selected_sheet]
    for idx, column in enumerate(column_names):
        print(f"列 {idx + 1}: {column}")

    keyword_column_index = int(input("请输入关键词列编号: ")) - 1
    target_column_index = int(input("请输入需要检查的列编号: ")) - 1

    # 确保选择的列存在
    df = pd.read_excel(excel_file, sheet_name=selected_sheet)
    if keyword_column_index < 0 or keyword_column_index >= len(column_names):
        print("所选择的关键词列编号无效。")
        return
    if target_column_index < 0 or target_column_index >= len(column_names):
        print("所选择的目标列编号无效。")
        return

    keyword_column = column_names[keyword_column_index]
    target_column = column_names[target_column_index]

    # 从指定的关键词列中获取非空值，并确保它们是字符串类型
    # df[keyword_column]：获取关键词列的数据
    keyword_series = df[keyword_column]  # 将关键词列的数据存储在变量中

    # 去掉空值
    non_null_keywords = keyword_series.dropna()  # 移除任何缺失值（NaN）

    # 将每个值转换为字符串
    string_keywords = non_null_keywords.astype(str)  # 确保所有值都是字符串

    # 将所有字符串连接为一个单一的字符串，以逗号分隔
    concatenated_keywords = string_keywords.str.cat(sep=',')  # 将所有字符串合并

    # 将连接后的字符串分割成列表
    keywords_list = concatenated_keywords.split(',')  # 以逗号为分隔符分割字符串

    # 去掉每个关键词的前后空格
    cleaned_keywords = [keyword.strip() for keyword in keywords_list]  # 去掉多余的空格

    # 结果是一个干净的关键词列表
    keywords = cleaned_keywords  # 最终的关键词列表

    # 加载工作簿
    workbook = load_workbook(excel_file)
    sheet = workbook[selected_sheet]

    # 定义红色背景
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 检查目标列中的每个单元格
    for row in range(2, sheet.max_row + 1):  # 从第二行开始（假设第一行为标题）
        cell_value = sheet.cell(row=row, column=target_column_index + 1).value  # 使用索引获取单元格
        if isinstance(cell_value, str):  # 确保单元格内容是字符串
            for keyword in keywords:
                if keyword in cell_value:
                    # 如果找到关键词，则标记单元格
                    sheet.cell(row=row, column=target_column_index + 1).fill = red_fill
                    break  # 一旦找到一个关键词就停止检查

    # 保存修改后的 Excel 文件
    workbook.save(excel_file)
    print("处理完成，Excel 文件已更新。")


if __name__ == "__main__":
    main()
