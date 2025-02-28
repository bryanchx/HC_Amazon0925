import openpyxl

def reorder_columns_in_template(source_file, output_file, new_column_order, sheet_name="Template"):
    # 读取源文件
    wb = openpyxl.load_workbook(source_file)
    print(f"成功加载源文件 '{source_file}'")

    # 获取指定的工作表
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表 '{sheet_name}' 在源文件中未找到")
    ws = wb[sheet_name]

    # 获取第二行的表头
    headers = [cell.value for cell in ws[2]]  # 第二行是表头
    print(f"当前表头: {headers}")  # 打印一下当前表头确认

    # 创建一个字典映射列名到列索引
    header_to_index = {header: idx + 1 for idx, header in enumerate(headers)}
    print(f"列名到列索引的映射: {header_to_index}")  # 打印列名到列索引的映射

    # 创建一个新的列索引顺序，依照指定的 new_column_order
    new_index_order = []

    for col in new_column_order:
        if col in header_to_index:
            new_index_order.append(header_to_index[col])
        else:
            print(f"列 '{col}' 在源文件中的表头中未找到，已跳过")  # 跳过列
            continue

    print(f"新的列顺序: {new_index_order}")  # 打印新的列顺序

    # 对列进行重排列
    for new_index, old_index in enumerate(new_index_order, start=1):
        if old_index != new_index:
            print(f"正在交换列 {headers[old_index - 1]}（旧列: {old_index}）和列 {headers[new_index - 1]}（新列: {new_index}）")

            # 获取当前列的数据
            col_cells = [cell.value for cell in next(ws.iter_cols(min_col=old_index, max_col=old_index, min_row=3))]
            print(f"列 {headers[old_index - 1]} 的数据: {col_cells[:5]}...")  # 打印前5个数据作为样本

            # 将数据插入到目标列位置
            for row_idx, value in enumerate(col_cells, start=3):  # 使用列中的数据进行填充
                ws.cell(row=row_idx, column=new_index, value=value)

            # 删除原列的数据
            for row in ws.iter_rows(min_row=3, min_col=old_index, max_col=old_index):
                for cell in row:
                    cell.value = None

            print(f"列 {headers[old_index - 1]} 的数据已移到列 {headers[new_index - 1]}")

    # 保存修改后的文件
    wb.save(output_file)
    print(f"列顺序已成功调整并保存为 '{output_file}'")

# 示例：指定列的顺序进行重排
source_file = r'E:\amazononline\汇丛\paEAI\CHX250103-PAEAI-Women-trousers-all-14-111本.xlsm'
output_file = r'output_file.xlsm'

# 假设这是你希望的列顺序
new_column_order = [
    'Product Type', 'Seller SKU', 'Parentage', 'Parent SKU', 'Product Name', 'Colour Map', 'Colour', 'Size',
    'Size Map', 'Bottoms Size Value', 'Outer Material Type', 'Your Price', 'List Price with Tax for Display',
    'Main Image Url', 'Other Image Url', 'Package Length', 'package-width', 'Package Height', 'Package Weight',
    'Product Description', 'Search Terms', 'Key Product Features', 'Recommended Browse Nodes', 'Brand Name',
    'Manufacturer', 'Language', 'Quantity', 'Target Gender', 'Age Range Description', 'Bottoms Size System',
    'Bottoms Size Class', 'Bottoms Body Type', 'Bottoms Height Type', 'Package Length Unit Of Measure',
    'Package Weight Unit Of Measure', 'Package Height Unit Of Measure', 'Package Width Unit Of Measure',
    'Fabric type', 'Product Care Instructions', 'Closure Type', 'Fur Description', 'Weave Type', 'Variation Theme',
    'Inner Material Type', 'Model Name', 'Model Number', 'Item Type', 'Season and collection year',
    'Occasion description',
    'Style Name', 'Fit Type', 'Special Features', 'size-modifier', 'platinum-keywords1 - platinum-keywords5',
    'Item Type Name', 'Occasion Type', 'Sport Type', 'Season', 'Athlete', 'Material Type', 'Rise Style', 'Front Style',
    'Pocket Description', 'Theme', 'League Name', 'Shaft Style Type', 'Product Lifecycle Supply Type', 'Department',
    'Is Autographed', 'Condition Type', 'Handling Time', 'Merchant Shipping Group', 'Release Date',
    'Country/Region Of Origin', 'Product ID', 'Product ID Type', 'Item Booking Date', 'Product Exemption Reason',
    'Temperature Rating', 'Manufacturer Part Number', 'Package Level', 'Package Contains Quantity',
    'Package Contains Identifier', 'Bottoms Size To Range', 'Waist Size Value', 'Inseam Size Value', 'Shipping Weight',
    'Website Shipping Weight Unit Of Measure', 'Item Width Unit Of Measure', 'Item Width', 'Item Height', 'Rise Height',
    'Item Shape', 'Item Height Unit Of Measure', 'Item Length Unit Of Measure', 'Item Length', 'Fulfillment Centre ID',
    'Legal Disclaimer Description', 'Safety Warning', 'EU Toys Safety Directive Age-specific warning',
    'Is this product a battery or does it utilise batteries?', 'Batteries are Included', 'Battery composition',
    'Battery type/size', 'Number of batteries', 'Battery weight (grams)', 'Number of Lithium Metal Cells',
    'Lithium Battery Packaging', 'Safety Data Sheet (SDS) URL', 'Item Weight', 'Volume', 'Validation Status',
    'Flash point (°C)', 'GPSR Safety Attestation', 'Material/Fabric Regulations', 'Condition Note', 'Currency',
    'Sale Price', 'Sale From Date', 'Sale End Date', 'Package Quantity', 'Number of Items', 'Can Be Gift Messaged',
    'Is Gift Wrap Available?', 'Launch Date', 'Restock Date', 'Minimum Advertised Price',
    'Fulfillment Center Shelf Life',
    'Offer End Date', 'Offer Start Date', 'Scheduled Delivery SKU List', 'RRP', 'Product Tax Code', 'Business Price',
    'Quantity Price Type', 'Quantity Lower Bound 1', 'Quantity Price 1', 'Quantity Lower Bound 2', 'Quantity Price 2',
    'Progressive Discount Type', 'Progressive Discount Lower Bound 1', 'Progressive Discount Value 1'
]

# 调用函数进行列顺序重排
reorder_columns_in_template(source_file, output_file, new_column_order)
